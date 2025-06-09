#now this my delete_admin.py backend file
from flask import request, redirect, url_for, send_file, session
from datetime import datetime, timedelta
from openpyxl import Workbook
import io
import math
from config import vehicles_collection, completed_records, rates_collection, users_collection, jobs_collection

def generate_report():
    if not session.get('is_admin'):
        return redirect(url_for('route_login'))
        
    report_type = request.form.get('report_type')
    date = request.form.get('date')
    selected_user = request.form.get('selected_user')
    
    # For delivery jobs, we allow any admin to generate reports for any user
    # Only verify that the user exists
    user = users_collection.find_one({'username': selected_user})
    if not user:
        return redirect(url_for('route_admin_dashboard', error='User not found'))
    
    try:
        date_obj = None
        next_date = None
        if date:  # Only parse date if provided
            date_obj = datetime.strptime(date, '%Y-%m-%d')
            next_date = date_obj + timedelta(days=1)
            
        workbook = Workbook()
        
        if report_type == 'current':
            # Pipeline for currently parked vehicles
            pipeline = [
                {
                    '$match': {
                        'checkout_time': None,
                        'handled_by': selected_user
                    }
                },
                {
                    '$project': {
                        'vehicle_number': 1,
                        'checkin_time': {
                            '$dateToString': {
                                'format': '%Y-%m-%d %H:%M:%S',
                                'date': '$checkin_time'
                            }
                        },
                        'payment_mode': 1,
                        'handled_by': 1,
                        '_id': 0
                    }
                }
            ]
            vehicles = list(vehicles_collection.aggregate(pipeline))
            filename = f'current_parked_vehicles_{selected_user}.xlsx'
            
        elif report_type == 'checkins':
            # Pipeline for check-ins by date
            pipeline = [
                {
                    '$match': {
                        'checkin_time': {
                            '$gte': date_obj,
                            '$lt': next_date
                        },
                        'handled_by': selected_user
                    }
                },
                {
                    '$project': {
                        'vehicle_number': 1,
                        'checkin_time': {
                            '$dateToString': {
                                'format': '%Y-%m-%d %H:%M:%S',
                                'date': '$checkin_time'
                            }
                        },
                        'payment_mode': 1,
                        'handled_by': 1,
                        '_id': 0
                    }
                }
            ]
            vehicles = list(vehicles_collection.aggregate(pipeline))
            filename = f'checkins_{selected_user}_{date}.xlsx'
            
        elif report_type == 'checkouts':
            # Pipeline for check-outs by date
            pipeline = [
                {
                    '$match': {
                        'checkout_time': {
                            '$gte': date_obj,
                            '$lt': next_date
                        },
                        'handled_by': selected_user
                    }
                },
                {
                    '$project': {
                        'vehicle_number': 1,
                        'checkin_time': {
                            '$dateToString': {
                                'format': '%Y-%m-%d %H:%M:%S',
                                'date': '$checkin_time'
                            }
                        },
                        'checkout_time': {
                            '$dateToString': {
                                'format': '%Y-%m-%d %H:%M:%S',
                                'date': '$checkout_time'
                            }
                        },
                        'payment_mode': 1,
                        'charge': 1,
                        'handled_by': 1,
                        '_id': 0
                    }
                }
            ]
            vehicles = list(completed_records.aggregate(pipeline))
            filename = f'checkouts_{selected_user}_{date}.xlsx'
            
        elif report_type == 'financial':
            # Financial report generation
            checkin_pipeline = [
                {
                    '$match': {
                        'checkin_time': {
                            '$gte': date_obj,
                            '$lt': next_date
                        },
                        'handled_by': selected_user
                    }
                },
                {
                    '$group': {
                        '_id': '$payment_mode',
                        'count': {'$sum': 1},
                        'total_amount': {'$sum': 15}
                    }
                }
            ]
            
            # Modified checkout pipeline to match actual database structure
            checkout_pipeline = [
                {
                    '$match': {
                        'checkout_time': {
                            '$gte': date_obj,
                            '$lt': next_date
                        },
                        'handled_by': selected_user
                    }
                },
                {
                    '$group': {
                        '_id': {
                            'initial_mode': '$initial_payment_mode',
                            'additional_mode': '$additional_payment_mode'
                        },
                        'count': {'$sum': 1},
                        'initial_amount': {'$sum': '$initial_payment'},
                        'additional_amount': {'$sum': '$additional_charge'},
                        'total_amount': {'$sum': '$total_charge'}
                    }
                }
            ]
            
            checkin_stats = list(vehicles_collection.aggregate(checkin_pipeline))
            checkout_stats = list(completed_records.aggregate(checkout_pipeline))
            
            # Create sheets
            checkin_sheet = workbook.active
            checkin_sheet.title = "Check-in Stats"
            
            # Write check-in stats
            checkin_sheet['A1'] = f"Check-in Statistics for {selected_user}"
            checkin_sheet['A2'] = "Payment Mode"
            checkin_sheet['B2'] = "Vehicle Count"
            checkin_sheet['C2'] = "Total Amount"
            
            row = 3
            total_checkins = 0
            total_checkin_amount = 0
            checkin_by_mode = {}
            
            for stat in checkin_stats:
                checkin_sheet[f'A{row}'] = stat['_id']
                checkin_sheet[f'B{row}'] = stat['count']
                checkin_sheet[f'C{row}'] = stat['total_amount']
                total_checkins += stat['count']
                total_checkin_amount += stat['total_amount']
                checkin_by_mode[stat['_id']] = stat['total_amount']
                row += 1
            
            checkin_sheet[f'A{row}'] = "TOTAL"
            checkin_sheet[f'B{row}'] = total_checkins
            checkin_sheet[f'C{row}'] = total_checkin_amount
            
            # Create modified check-out sheet with payment mode splits
            checkout_sheet = workbook.create_sheet("Check-out Stats")
            checkout_sheet['A1'] = f"Check-out Statistics for {selected_user}"
            checkout_sheet['A2'] = "Payment Mode"
            checkout_sheet['B2'] = "Vehicle Count"
            checkout_sheet['C2'] = "Initial Amount"
            checkout_sheet['D2'] = "Additional Amount"
            checkout_sheet['E2'] = "Total Amount"
            
            row = 3
            total_checkouts = 0
            total_initial_amount = 0
            total_additional_amount = 0
            total_checkout_amount = 0
            checkout_summary = {}  # To store amounts by payment mode
            
            for stat in checkout_stats:
                initial_mode = stat['_id']['initial_mode']
                additional_mode = stat['_id']['additional_mode']
                
                # Handle initial payment
                if initial_mode not in checkout_summary:
                    checkout_summary[initial_mode] = {
                        'count': 0,
                        'initial': 0,
                        'additional': 0,
                        'total': 0
                    }
                checkout_summary[initial_mode]['count'] += stat['count']
                checkout_summary[initial_mode]['initial'] += stat['initial_amount']
                checkout_summary[initial_mode]['total'] += stat['initial_amount']
                
                # Handle additional payment if exists
                if additional_mode:
                    if additional_mode not in checkout_summary:
                        checkout_summary[additional_mode] = {
                            'count': 0,
                            'initial': 0,
                            'additional': 0,
                            'total': 0
                        }
                    checkout_summary[additional_mode]['additional'] += stat['additional_amount']
                    checkout_summary[additional_mode]['total'] += stat['additional_amount']
                
                total_checkouts += stat['count']
                total_initial_amount += stat['initial_amount']
                total_additional_amount += stat['additional_amount']
                total_checkout_amount += stat['total_amount']
            
            # Write checkout summary by payment mode
            for mode, data in checkout_summary.items():
                if mode:  # Skip empty payment modes
                    checkout_sheet[f'A{row}'] = mode
                    checkout_sheet[f'B{row}'] = data['count']
                    checkout_sheet[f'C{row}'] = data['initial']
                    checkout_sheet[f'D{row}'] = data['additional']
                    checkout_sheet[f'E{row}'] = data['total']
                    row += 1
            
            # Write totals
            checkout_sheet[f'A{row}'] = "TOTAL"
            checkout_sheet[f'B{row}'] = total_checkouts
            checkout_sheet[f'C{row}'] = total_initial_amount
            checkout_sheet[f'D{row}'] = total_additional_amount
            checkout_sheet[f'E{row}'] = total_checkout_amount
            
            # Create detailed summary sheet
            summary_sheet = workbook.create_sheet("Daily Summary")
            summary_sheet['A1'] = f"Financial Summary for {selected_user} on {date}"
            
            row = 3
            # Check-in Summary
            summary_sheet[f'A{row}'] = "CHECK-IN SUMMARY"
            row += 1
            summary_sheet[f'A{row}'] = "Total Vehicles Checked In"
            summary_sheet[f'B{row}'] = total_checkins
            row += 1
            
            summary_sheet[f'A{row}'] = "Check-in Amount by Payment Mode:"
            row += 1
            for mode, amount in checkin_by_mode.items():
                summary_sheet[f'A{row}'] = f"Total Check-in Amount ({mode})"
                summary_sheet[f'B{row}'] = amount
                row += 1
            
            summary_sheet[f'A{row}'] = "Total Check-in Amount"
            summary_sheet[f'B{row}'] = total_checkin_amount
            row += 2
            
            # Check-out Summary
            summary_sheet[f'A{row}'] = "CHECK-OUT SUMMARY"
            row += 1
            summary_sheet[f'A{row}'] = "Total Vehicles Checked Out"
            summary_sheet[f'B{row}'] = total_checkouts
            row += 1
            
            # Initial Amount Summary
            summary_sheet[f'A{row}'] = "Initial Amount by Payment Mode:"
            row += 1
            for mode, data in checkout_summary.items():
                if mode and data['initial'] > 0:
                    summary_sheet[f'A{row}'] = f"Initial Amount ({mode})"
                    summary_sheet[f'B{row}'] = data['initial']
                    row += 1
            
            summary_sheet[f'A{row}'] = "Total Initial Amount"
            summary_sheet[f'B{row}'] = total_initial_amount
            row += 2
            
            # Additional Amount Summary
            summary_sheet[f'A{row}'] = "Additional Amount by Payment Mode:"
            row += 1
            for mode, data in checkout_summary.items():
                if mode and data['additional'] > 0:
                    summary_sheet[f'A{row}'] = f"Additional Amount ({mode})"
                    summary_sheet[f'B{row}'] = data['additional']
                    row += 1
            
            summary_sheet[f'A{row}'] = "Total Additional Amount"
            summary_sheet[f'B{row}'] = total_additional_amount
            row += 2
            
            # Total Business Summary
            summary_sheet[f'A{row}'] = "TOTAL BUSINESS SUMMARY"
            row += 1
            
            # Combine check-in and check-out amounts by payment mode
            total_by_mode = {}
            for mode, amount in checkin_by_mode.items():
                if mode not in total_by_mode:
                    total_by_mode[mode] = 0
                total_by_mode[mode] += amount
            
            for mode, data in checkout_summary.items():
                if mode not in total_by_mode:
                    total_by_mode[mode] = 0
                total_by_mode[mode] += data['total']
            
            for mode, total in total_by_mode.items():
                if mode:  # Skip empty payment modes
                    summary_sheet[f'A{row}'] = f"Total Business ({mode})"
                    summary_sheet[f'B{row}'] = total
                    row += 1
            
            summary_sheet[f'A{row}'] = "TOTAL BUSINESS FOR THE DAY"
            summary_sheet[f'B{row}'] = total_checkin_amount + total_checkout_amount
            
            filename = f'financial_report_{selected_user}_{date}.xlsx'
        
        # Add new delivery job report type
        elif report_type == 'delivery_jobs':
            # Create a sheet for delivery jobs report
            jobs_sheet = workbook.active
            jobs_sheet.title = "Delivery Jobs"
            
            # Setup headers
            jobs_sheet['A1'] = f"Delivery Jobs Report for {selected_user}"
            jobs_sheet['A3'] = "Job ID"
            jobs_sheet['B3'] = "From Location"
            jobs_sheet['C3'] = "To Location"
            jobs_sheet['D3'] = "Created At"
            jobs_sheet['E3'] = "Deadline"
            jobs_sheet['F3'] = "Status"
            jobs_sheet['G3'] = "Details"
            
            # Get delivery jobs data based on the selected user
            if date_obj:
                # Filter by date if provided
                jobs_query = {
                    '$or': [
                        {'assigned_to': selected_user},
                        {'created_by': selected_user}
                    ],
                    'created_at': {
                        '$gte': date_obj,
                        '$lt': next_date
                    }
                }
            else:
                # Get all jobs for the user
                jobs_query = {
                    '$or': [
                        {'assigned_to': selected_user},
                        {'created_by': selected_user}
                    ]
                }
            
            delivery_jobs = list(jobs_collection.find(jobs_query).sort("created_at", -1))
            
            # Initialize vehicles as empty to avoid NameError later
            vehicles = []
            
            # If no jobs found, add a message
            if not delivery_jobs:
                jobs_sheet['A4'] = "No delivery jobs found for this user/date."
            else:
                # Fill the sheet with job data
                row = 4
                for job in delivery_jobs:
                    jobs_sheet[f'A{row}'] = job.get('job_id', str(job.get('_id', '')))
                    jobs_sheet[f'B{row}'] = job.get('load_from', '')
                    jobs_sheet[f'C{row}'] = job.get('load_to', '')
                    jobs_sheet[f'D{row}'] = job.get('created_at').strftime('%Y-%m-%d %H:%M:%S') if job.get('created_at') else 'N/A'
                    jobs_sheet[f'E{row}'] = job.get('deadline').strftime('%Y-%m-%d %H:%M:%S') if job.get('deadline') else 'N/A'
                    jobs_sheet[f'F{row}'] = job.get('status', 'Unknown').capitalize().replace('_', ' ')
                    jobs_sheet[f'G{row}'] = job.get('load_details', '')
                    row += 1
                
                # Add summary statistics
                row += 2
                jobs_sheet[f'A{row}'] = "Summary Statistics:"
                row += 1
                
                # Count jobs by status
                status_counts = {}
                for job in delivery_jobs:
                    status = job.get('status', 'unknown')
                    status_counts[status] = status_counts.get(status, 0) + 1
                
                jobs_sheet[f'A{row}'] = "Status"
                jobs_sheet[f'B{row}'] = "Count"
                row += 1
                
                for status, count in status_counts.items():
                    jobs_sheet[f'A{row}'] = status.capitalize().replace('_', ' ')
                    jobs_sheet[f'B{row}'] = count
                    row += 1
                
                row += 1
                jobs_sheet[f'A{row}'] = "Total Jobs:"
                jobs_sheet[f'B{row}'] = len(delivery_jobs)
            
            # Add a performance summary sheet
            performance_sheet = workbook.create_sheet("Performance Summary")
            
            performance_sheet['A1'] = f"Delivery Job Performance for {selected_user}"
            performance_sheet['A3'] = "Metric"
            performance_sheet['B3'] = "Value"
            
            row = 4
            
            # Calculate performance metrics
            total_jobs = len(delivery_jobs)
            completed_jobs = sum(1 for job in delivery_jobs if job.get('status') == 'completed')
            in_progress_jobs = sum(1 for job in delivery_jobs if job.get('status') == 'in_progress')
            open_jobs = sum(1 for job in delivery_jobs if job.get('status') == 'open')
            assigned_jobs = sum(1 for job in delivery_jobs if job.get('status') == 'assigned')
            
            # Calculate on-time delivery rate
            on_time_jobs = 0
            late_jobs = 0
            jobs_with_deadline = 0
            
            for job in delivery_jobs:
                if job.get('status') == 'completed' and job.get('deadline'):
                    jobs_with_deadline += 1
                    if job.get('updated_at', job.get('created_at')) <= job.get('deadline'):
                        on_time_jobs += 1
                    else:
                        late_jobs += 1
            
            on_time_percentage = (on_time_jobs / jobs_with_deadline * 100) if jobs_with_deadline > 0 else 0
            
            # Add metrics to the sheet
            performance_sheet[f'A{row}'] = "Total Jobs"
            performance_sheet[f'B{row}'] = total_jobs
            row += 1
            
            performance_sheet[f'A{row}'] = "Completed Jobs"
            performance_sheet[f'B{row}'] = completed_jobs
            row += 1
            
            performance_sheet[f'A{row}'] = "In Progress Jobs"
            performance_sheet[f'B{row}'] = in_progress_jobs
            row += 1
            
            performance_sheet[f'A{row}'] = "Assigned Jobs"
            performance_sheet[f'B{row}'] = assigned_jobs
            row += 1
            
            performance_sheet[f'A{row}'] = "Open Jobs"
            performance_sheet[f'B{row}'] = open_jobs
            row += 1
            
            performance_sheet[f'A{row}'] = "Completion Rate"
            performance_sheet[f'B{row}'] = f"{(completed_jobs / total_jobs * 100) if total_jobs > 0 else 0:.2f}%"
            row += 1
            
            performance_sheet[f'A{row}'] = "On-time Delivery Rate"
            performance_sheet[f'B{row}'] = f"{on_time_percentage:.2f}%"
            
            filename = f'delivery_jobs_{selected_user}_{date if date else "all_time"}.xlsx'
        
        # Write to Excel for non-financial reports
        if report_type != 'financial' and report_type != 'delivery_jobs':
            worksheet = workbook.active
            if vehicles:
                headers = list(vehicles[0].keys())
                for col, header in enumerate(headers, start=1):
                    worksheet.cell(row=1, column=col, value=header.replace('_', ' ').title())
                
                for row, vehicle in enumerate(vehicles, start=2):
                    for col, key in enumerate(headers, start=1):
                        worksheet.cell(row=row, column=col, value=vehicle[key])
        
        # Create response
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"Error generating report: {str(e)}")
        return redirect(url_for('route_admin_dashboard', error='Error generating report')) 